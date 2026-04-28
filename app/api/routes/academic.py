from __future__ import annotations

from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from sqlalchemy import distinct, func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.admin_auth import require_admin_key
from app.api.deps import get_db
from app.models.academic import Faculty, Specialty, University
from app.schemas.academic import (
    AcademicImportOut,
    FacultyIn,
    FacultyOut,
    FacultyUpdate,
    SpecialtyIn,
    SpecialtyListOut,
    SpecialtyOut,
    SpecialtyPageOut,
    SpecialtyStatsOut,
    SpecialtyUpdate,
    UniversityIn,
    UniversityOut,
    UniversityUpdate,
)
from app.services.academic_import_service import AcademicImportService, build_specialty_entry_key

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - environment-dependent
    load_workbook = None


admin_router = APIRouter(prefix="/admin/academic", tags=["admin-academic"], dependencies=[Depends(require_admin_key)])
public_router = APIRouter(prefix="/academic", tags=["academic"])


def _to_text(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _university_to_out(row: University) -> UniversityOut:
    return UniversityOut(
        id=row.id,
        serial_no=row.serial_no,
        name=row.name,
        region=row.region,
        city=row.city,
        district=row.district,
        phone=row.phone,
    )


def _faculty_to_out(row: Faculty) -> FacultyOut:
    university_name = row.university.name if row.university is not None else None
    return FacultyOut(
        id=row.id,
        university_id=row.university_id,
        code=row.code,
        name=row.name,
        source_sheet=row.source_sheet,
        university_name=university_name,
    )


def _specialty_to_out(row: Specialty) -> SpecialtyOut:
    return SpecialtyOut(
        id=row.id,
        faculty_id=row.faculty_id,
        excel_id=row.excel_id,
        code=row.code,
        name=row.name,
        study_mode=row.study_mode,
        language=row.language,
        tuition=row.tuition,
        admission_quota=row.admission_quota,
        degree=row.degree,
        is_free=row.is_free,
        price=row.price,
        source_sheet=row.source_sheet,
    )


@admin_router.get("/universities", response_model=list[UniversityOut])
def list_universities(db: Session = Depends(get_db)):
    rows = db.execute(select(University).order_by(University.name, University.id)).scalars().all()
    return [_university_to_out(r) for r in rows]


@admin_router.post("/universities", response_model=UniversityOut)
def create_university(body: UniversityIn, db: Session = Depends(get_db)):
    row = University(
        serial_no=body.serial_no,
        name=body.name.strip(),
        region=_to_text(body.region),
        city=_to_text(body.city),
        district=_to_text(body.district),
        phone=_to_text(body.phone),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _university_to_out(row)


@admin_router.patch("/universities/{university_id}", response_model=UniversityOut)
def patch_university(university_id: int, body: UniversityUpdate, db: Session = Depends(get_db)):
    row = db.get(University, university_id)
    if not row:
        raise HTTPException(status_code=404, detail="University not found")
    data = body.model_dump(exclude_unset=True)
    if "serial_no" in data:
        row.serial_no = data["serial_no"]
    if "name" in data and data["name"] is not None:
        row.name = data["name"].strip()
    if "region" in data:
        row.region = _to_text(data["region"])
    if "city" in data:
        row.city = _to_text(data["city"])
    if "district" in data:
        row.district = _to_text(data["district"])
    if "phone" in data:
        row.phone = _to_text(data["phone"])
    db.commit()
    db.refresh(row)
    return _university_to_out(row)


@admin_router.delete("/universities/{university_id}", status_code=204)
def delete_university(university_id: int, db: Session = Depends(get_db)):
    row = db.get(University, university_id)
    if not row:
        raise HTTPException(status_code=404, detail="University not found")
    db.delete(row)
    db.commit()


@admin_router.get("/faculties", response_model=list[FacultyOut])
def list_faculties(
    university_id: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
):
    stmt = select(Faculty, University.name).join(University, University.id == Faculty.university_id)
    if university_id is not None:
        stmt = stmt.where(Faculty.university_id == university_id)
    stmt = stmt.order_by(Faculty.university_id, Faculty.code, Faculty.name, Faculty.id)
    rows = db.execute(stmt).all()
    return [
        FacultyOut(
            id=faculty.id,
            university_id=faculty.university_id,
            code=faculty.code,
            name=faculty.name,
            source_sheet=faculty.source_sheet,
            university_name=university_name,
        )
        for faculty, university_name in rows
    ]


@admin_router.post("/faculties", response_model=FacultyOut)
def create_faculty(body: FacultyIn, db: Session = Depends(get_db)):
    university = db.get(University, body.university_id)
    if not university:
        raise HTTPException(status_code=404, detail="University not found")
    row = Faculty(
        university_id=body.university_id,
        code=_to_text(body.code),
        name=body.name.strip(),
        source_sheet=_to_text(body.source_sheet),
    )
    db.add(row)
    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail="Faculty already exists in this university") from e
    db.refresh(row)
    row.university = university
    return _faculty_to_out(row)


@admin_router.patch("/faculties/{faculty_id}", response_model=FacultyOut)
def patch_faculty(faculty_id: int, body: FacultyUpdate, db: Session = Depends(get_db)):
    row = db.get(Faculty, faculty_id)
    if not row:
        raise HTTPException(status_code=404, detail="Faculty not found")
    data = body.model_dump(exclude_unset=True)
    if "university_id" in data and data["university_id"] is not None:
        if not db.get(University, data["university_id"]):
            raise HTTPException(status_code=404, detail="University not found")
        row.university_id = data["university_id"]
    if "code" in data:
        row.code = _to_text(data["code"])
    if "name" in data and data["name"] is not None:
        row.name = data["name"].strip()
    if "source_sheet" in data:
        row.source_sheet = _to_text(data["source_sheet"])
    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail="Faculty already exists in this university") from e
    db.refresh(row)
    row.university = db.get(University, row.university_id)
    return _faculty_to_out(row)


@admin_router.delete("/faculties/{faculty_id}", status_code=204)
def delete_faculty(faculty_id: int, db: Session = Depends(get_db)):
    row = db.get(Faculty, faculty_id)
    if not row:
        raise HTTPException(status_code=404, detail="Faculty not found")
    db.delete(row)
    db.commit()


def _specialty_row_to_out(
    row: Specialty,
    faculty_name: str,
    faculty_code: Optional[str],
    university_id: int,
    university_name: str,
    region: Optional[str],
    city: Optional[str],
    district: Optional[str],
    phone: Optional[str],
) -> SpecialtyListOut:
    return SpecialtyListOut(
        id=row.id,
        faculty_id=row.faculty_id,
        excel_id=row.excel_id,
        code=row.code,
        name=row.name,
        study_mode=row.study_mode,
        language=row.language,
        tuition=row.tuition,
        admission_quota=row.admission_quota,
        degree=row.degree,
        is_free=row.is_free,
        price=row.price,
        source_sheet=row.source_sheet,
        faculty_name=faculty_name,
        faculty_code=faculty_code,
        university_id=university_id,
        university_name=university_name,
        region=region,
        city=city,
        district=district,
        phone=phone,
    )


@admin_router.get("/specialties", response_model=list[SpecialtyListOut])
@public_router.get("/specialties", response_model=list[SpecialtyListOut])
def list_specialties(
    university_id: Optional[int] = Query(default=None),
    faculty_id: Optional[int] = Query(default=None),
    specialty_id: Optional[int] = Query(default=None),
    group_code: Optional[str] = Query(default=None, description="Код группы / лист Excel"),
    samt: Optional[str] = Query(default=None, description="Группа / факультет / самт — частичное совпадение"),
    university: Optional[str] = Query(default=None, description="Название муассиса — частичное совпадение"),
    makon: Optional[str] = Query(default=None, description="Город или район / колонка макон"),
    code_name: Optional[str] = Query(default=None, description="Рамз ё номи ихтисос"),
    study_mode: Optional[str] = Query(default=None),
    tuition: Optional[str] = Query(default=None, description="Намуди таҳсил / маблағ"),
    language: Optional[str] = Query(default=None),
    admission_quota: Optional[str] = Query(default=None, description="Нақшаи қабул"),
    degree: Optional[str] = Query(default=None),
    free_only: Optional[bool] = Query(default=None),
    price_min: Optional[int] = Query(default=None, ge=0),
    price_max: Optional[int] = Query(default=None, ge=0),
    q: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    stmt = (
        select(
            Specialty,
            Faculty.name,
            Faculty.code,
            University.id,
            University.name,
            University.region,
            University.city,
            University.district,
            University.phone,
        )
        .join(Faculty, Faculty.id == Specialty.faculty_id)
        .join(University, University.id == Faculty.university_id)
    )
    if specialty_id is not None:
        stmt = stmt.where(Specialty.id == specialty_id)
    if university_id is not None:
        stmt = stmt.where(University.id == university_id)
    if faculty_id is not None:
        stmt = stmt.where(Faculty.id == faculty_id)
    if group_code and group_code.strip():
        term = f"%{group_code.strip().lower()}%"
        stmt = stmt.where(func.lower(func.coalesce(Faculty.code, "")).like(term))
    if samt and samt.strip():
        term = f"%{samt.strip().lower()}%"
        stmt = stmt.where(
            or_(
                func.lower(Faculty.name).like(term),
                func.lower(func.coalesce(Faculty.code, "")).like(term),
            )
        )
    if university and university.strip():
        term = f"%{university.strip().lower()}%"
        stmt = stmt.where(func.lower(University.name).like(term))
    if makon and makon.strip():
        term = f"%{makon.strip().lower()}%"
        stmt = stmt.where(
            or_(
                func.lower(func.coalesce(University.city, "")).like(term),
                func.lower(func.coalesce(University.district, "")).like(term),
                func.lower(
                    func.trim(
                        func.concat_ws(
                            " ",
                            func.coalesce(University.city, ""),
                            func.coalesce(University.district, ""),
                        )
                    )
                ).like(term),
            )
        )
    if code_name and code_name.strip():
        term = f"%{code_name.strip().lower()}%"
        stmt = stmt.where(
            func.lower(Specialty.name).like(term) | func.lower(func.coalesce(Specialty.code, "")).like(term)
        )
    if study_mode and study_mode.strip():
        term = f"%{study_mode.strip().lower()}%"
        stmt = stmt.where(func.lower(func.coalesce(Specialty.study_mode, "")).like(term))
    if tuition and tuition.strip():
        term = f"%{tuition.strip().lower()}%"
        stmt = stmt.where(func.lower(func.coalesce(Specialty.tuition, "")).like(term))
    if language and language.strip():
        term = f"%{language.strip().lower()}%"
        stmt = stmt.where(func.lower(func.coalesce(Specialty.language, "")).like(term))
    if admission_quota and admission_quota.strip():
        term = f"%{admission_quota.strip().lower()}%"
        stmt = stmt.where(func.lower(func.coalesce(Specialty.admission_quota, "")).like(term))
    if degree and degree.strip():
        term = f"%{degree.strip().lower()}%"
        stmt = stmt.where(func.lower(func.coalesce(Specialty.degree, "")).like(term))
    if free_only is not None:
        stmt = stmt.where(Specialty.is_free.is_(free_only))
    if price_min is not None:
        stmt = stmt.where(Specialty.price >= price_min)
    if price_max is not None:
        stmt = stmt.where(Specialty.price <= price_max)
    if q:
        term = f"%{q.strip().lower()}%"
        stmt = stmt.where(
            func.lower(Specialty.name).like(term)
            | func.lower(func.coalesce(Specialty.code, "")).like(term)
            | func.lower(func.coalesce(Specialty.degree, "")).like(term)
            | func.lower(Faculty.name).like(term)
            | func.lower(func.coalesce(Faculty.code, "")).like(term)
            | func.lower(University.name).like(term)
        )
    stmt = stmt.order_by(University.name, Faculty.code, Faculty.name, Specialty.name, Specialty.id)
    rows = db.execute(stmt).all()
    return [_specialty_row_to_out(*r) for r in rows]


@admin_router.get("/specialties/page", response_model=SpecialtyPageOut)
@public_router.get("/specialties/page", response_model=SpecialtyPageOut)
def list_specialties_page(
    university_id: Optional[int] = Query(default=None),
    faculty_id: Optional[int] = Query(default=None),
    specialty_id: Optional[int] = Query(default=None),
    group_code: Optional[str] = Query(default=None, description="Код группы / лист Excel"),
    samt: Optional[str] = Query(default=None, description="Группа / факультет / самт — частичное совпадение"),
    university: Optional[str] = Query(default=None, description="Название муассиса — частичное совпадение"),
    makon: Optional[str] = Query(default=None, description="Город или район / колонка макон"),
    code_name: Optional[str] = Query(default=None, description="Рамз ё номи ихтисос"),
    study_mode: Optional[str] = Query(default=None),
    tuition: Optional[str] = Query(default=None, description="Намуди таҳсил / маблағ"),
    language: Optional[str] = Query(default=None),
    admission_quota: Optional[str] = Query(default=None, description="Нақшаи қабул"),
    degree: Optional[str] = Query(default=None),
    free_only: Optional[bool] = Query(default=None),
    price_min: Optional[int] = Query(default=None, ge=0),
    price_max: Optional[int] = Query(default=None, ge=0),
    q: Optional[str] = Query(default=None),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, ge=1, le=500),
    db: Session = Depends(get_db),
):
    stmt = (
        select(
            Specialty,
            Faculty.name,
            Faculty.code,
            University.id,
            University.name,
            University.region,
            University.city,
            University.district,
            University.phone,
        )
        .join(Faculty, Faculty.id == Specialty.faculty_id)
        .join(University, University.id == Faculty.university_id)
    )
    if specialty_id is not None:
        stmt = stmt.where(Specialty.id == specialty_id)
    if university_id is not None:
        stmt = stmt.where(University.id == university_id)
    if faculty_id is not None:
        stmt = stmt.where(Faculty.id == faculty_id)
    if group_code and group_code.strip():
        term = f"%{group_code.strip().lower()}%"
        stmt = stmt.where(func.lower(func.coalesce(Faculty.code, "")).like(term))
    if samt and samt.strip():
        term = f"%{samt.strip().lower()}%"
        stmt = stmt.where(
            or_(
                func.lower(Faculty.name).like(term),
                func.lower(func.coalesce(Faculty.code, "")).like(term),
            )
        )
    if university and university.strip():
        term = f"%{university.strip().lower()}%"
        stmt = stmt.where(func.lower(University.name).like(term))
    if makon and makon.strip():
        term = f"%{makon.strip().lower()}%"
        stmt = stmt.where(
            or_(
                func.lower(func.coalesce(University.city, "")).like(term),
                func.lower(func.coalesce(University.district, "")).like(term),
                func.lower(
                    func.trim(
                        func.concat_ws(
                            " ",
                            func.coalesce(University.city, ""),
                            func.coalesce(University.district, ""),
                        )
                    )
                ).like(term),
            )
        )
    if code_name and code_name.strip():
        term = f"%{code_name.strip().lower()}%"
        stmt = stmt.where(
            func.lower(Specialty.name).like(term) | func.lower(func.coalesce(Specialty.code, "")).like(term)
        )
    if study_mode and study_mode.strip():
        term = f"%{study_mode.strip().lower()}%"
        stmt = stmt.where(func.lower(func.coalesce(Specialty.study_mode, "")).like(term))
    if tuition and tuition.strip():
        term = f"%{tuition.strip().lower()}%"
        stmt = stmt.where(func.lower(func.coalesce(Specialty.tuition, "")).like(term))
    if language and language.strip():
        term = f"%{language.strip().lower()}%"
        stmt = stmt.where(func.lower(func.coalesce(Specialty.language, "")).like(term))
    if admission_quota and admission_quota.strip():
        term = f"%{admission_quota.strip().lower()}%"
        stmt = stmt.where(func.lower(func.coalesce(Specialty.admission_quota, "")).like(term))
    if degree and degree.strip():
        term = f"%{degree.strip().lower()}%"
        stmt = stmt.where(func.lower(func.coalesce(Specialty.degree, "")).like(term))
    if free_only is not None:
        stmt = stmt.where(Specialty.is_free.is_(free_only))
    if price_min is not None:
        stmt = stmt.where(Specialty.price >= price_min)
    if price_max is not None:
        stmt = stmt.where(Specialty.price <= price_max)
    if q:
        term = f"%{q.strip().lower()}%"
        stmt = stmt.where(
            func.lower(Specialty.name).like(term)
            | func.lower(func.coalesce(Specialty.code, "")).like(term)
            | func.lower(func.coalesce(Specialty.degree, "")).like(term)
            | func.lower(Faculty.name).like(term)
            | func.lower(func.coalesce(Faculty.code, "")).like(term)
            | func.lower(University.name).like(term)
        )

    total = db.scalar(select(func.count()).select_from(stmt.order_by(None).subquery())) or 0
    total_pages = (total + limit - 1) // limit if total else 0
    rows = db.execute(
        stmt.order_by(University.name, Faculty.code, Faculty.name, Specialty.name, Specialty.id)
        .offset((page - 1) * limit)
        .limit(limit)
    ).all()
    return SpecialtyPageOut(
        data=[_specialty_row_to_out(*r) for r in rows],
        page=page,
        limit=limit,
        total=total,
        total_pages=total_pages,
    )


@admin_router.get("/specialties/stats", response_model=SpecialtyStatsOut)
@public_router.get("/specialties/stats", response_model=SpecialtyStatsOut)
def specialties_stats(db: Session = Depends(get_db)):
    total = db.scalar(select(func.count(Specialty.id))) or 0
    universities_count = (
        db.scalar(
            select(func.count(distinct(University.id)))
            .select_from(Specialty)
            .join(Faculty, Faculty.id == Specialty.faculty_id)
            .join(University, University.id == Faculty.university_id)
        )
        or 0
    )
    faculties_count = (
        db.scalar(
            select(func.count(distinct(Faculty.id)))
            .select_from(Specialty)
            .join(Faculty, Faculty.id == Specialty.faculty_id)
        )
        or 0
    )
    languages_count = (
        db.scalar(
            select(func.count(distinct(Specialty.language))).where(
                Specialty.language.is_not(None),
                Specialty.language != "",
            )
        )
        or 0
    )
    study_modes_count = (
        db.scalar(
            select(func.count(distinct(Specialty.study_mode))).where(
                Specialty.study_mode.is_not(None),
                Specialty.study_mode != "",
            )
        )
        or 0
    )
    free_count = db.scalar(select(func.count(Specialty.id)).where(Specialty.is_free.is_(True))) or 0
    paid_count = db.scalar(select(func.count(Specialty.id)).where(Specialty.is_free.is_(False))) or 0
    return SpecialtyStatsOut(
        total_specialties=total,
        universities_count=universities_count,
        faculties_count=faculties_count,
        languages_count=languages_count,
        study_modes_count=study_modes_count,
        free_count=free_count,
        paid_count=paid_count,
    )


@admin_router.post("/specialties", response_model=SpecialtyOut)
def create_specialty(body: SpecialtyIn, db: Session = Depends(get_db)):
    if not db.get(Faculty, body.faculty_id):
        raise HTTPException(status_code=404, detail="Faculty not found")
    entry_key = build_specialty_entry_key(
        excel_id=body.excel_id,
        faculty_id=body.faculty_id,
        code=_to_text(body.code),
        name=body.name.strip(),
        study_mode=_to_text(body.study_mode),
        language=_to_text(body.language),
        tuition=_to_text(body.tuition),
        admission_quota=_to_text(body.admission_quota),
        degree=_to_text(body.degree),
    )
    row = Specialty(
        faculty_id=body.faculty_id,
        excel_id=body.excel_id,
        entry_key=entry_key,
        code=_to_text(body.code),
        name=body.name.strip(),
        study_mode=_to_text(body.study_mode),
        language=_to_text(body.language),
        tuition=_to_text(body.tuition),
        admission_quota=_to_text(body.admission_quota),
        degree=_to_text(body.degree),
        is_free=body.is_free,
        price=body.price,
        source_sheet=_to_text(body.source_sheet),
    )
    db.add(row)
    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail="Specialty entry already exists") from e
    db.refresh(row)
    return _specialty_to_out(row)


@admin_router.patch("/specialties/{specialty_id}", response_model=SpecialtyOut)
def patch_specialty(specialty_id: int, body: SpecialtyUpdate, db: Session = Depends(get_db)):
    row = db.get(Specialty, specialty_id)
    if not row:
        raise HTTPException(status_code=404, detail="Specialty not found")
    data = body.model_dump(exclude_unset=True)
    if "faculty_id" in data and data["faculty_id"] is not None:
        if not db.get(Faculty, data["faculty_id"]):
            raise HTTPException(status_code=404, detail="Faculty not found")
        row.faculty_id = data["faculty_id"]
    if "excel_id" in data:
        row.excel_id = data["excel_id"]
    if "code" in data:
        row.code = _to_text(data["code"])
    if "name" in data and data["name"] is not None:
        row.name = data["name"].strip()
    if "study_mode" in data:
        row.study_mode = _to_text(data["study_mode"])
    if "language" in data:
        row.language = _to_text(data["language"])
    if "tuition" in data:
        row.tuition = _to_text(data["tuition"])
    if "admission_quota" in data:
        row.admission_quota = _to_text(data["admission_quota"])
    if "degree" in data:
        row.degree = _to_text(data["degree"])
    if "is_free" in data:
        row.is_free = data["is_free"]
    if "price" in data:
        row.price = data["price"]
    if "source_sheet" in data:
        row.source_sheet = _to_text(data["source_sheet"])

    row.entry_key = build_specialty_entry_key(
        excel_id=row.excel_id,
        faculty_id=row.faculty_id,
        code=row.code,
        name=row.name,
        study_mode=row.study_mode,
        language=row.language,
        tuition=row.tuition,
        admission_quota=row.admission_quota,
        degree=row.degree,
    )
    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail="Specialty entry already exists") from e
    db.refresh(row)
    return _specialty_to_out(row)


@admin_router.delete("/specialties/{specialty_id}", status_code=204)
def delete_specialty(specialty_id: int, db: Session = Depends(get_db)):
    row = db.get(Specialty, specialty_id)
    if not row:
        raise HTTPException(status_code=404, detail="Specialty not found")
    db.delete(row)
    db.commit()


@admin_router.post("/import", response_model=AcademicImportOut)
async def import_academic_excel(
    file: UploadFile = File(...),
    clear_existing: bool = Form(default=False),
    db: Session = Depends(get_db),
):
    if load_workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl is required for Excel import")
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    raw = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:  # pragma: no cover - input dependent
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}") from e

    stats = AcademicImportService(db).import_workbook(wb, clear_existing=clear_existing)
    db.commit()
    return AcademicImportOut(**stats)
